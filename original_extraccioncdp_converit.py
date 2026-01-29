import re
import fitz  # PyMuPDF
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

def limpiar_numero(s: str) -> int:
    if not s or s in ["-", ""]:
        return 0
    s = str(s).replace("$", "").replace(" ", "").strip()
    s = s.replace(".", "").replace(",", "")
    return int(s) if s.isdigit() else 0

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    texto = str(texto).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto

def convertir_pep(numero: str) -> str:
    numero = numero.zfill(5)
    return f"PM/0005/0101/4599000{numero}"

def extraer_cdps(pdf_path, log_lines):
    doc = fitz.open(pdf_path)
    lineas = []
    for page in doc:
        lineas.extend(page.get_text("text").splitlines())

    valor = 0
    objeto = ""
    numero_proyecto = None
    numero_oficio = "No encontrado"
    fecha_oficio = datetime.today().strftime("%d/%m/%Y")

    for idx, line in enumerate(lineas):
        # VALOR: número en la siguiente línea
        if "VALOR" in line.upper():
            if idx+1 < len(lineas):
                valor_line = lineas[idx+1]
                valor_match = re.search(r"([\d\.,]+)", valor_line)
                if valor_match:
                    valor = limpiar_numero(valor_match.group(1))

        # OBJETO: concatenar hasta encontrar VALOR
        if "OBJETO" in line.upper():
            objeto_lines = []
            for j in range(idx+1, len(lineas)):
                if "VALOR" in lineas[j].upper():
                    break
                objeto_lines.append(lineas[j])
            objeto = normalizar_texto(" ".join(objeto_lines))

        # Proyecto: buscar número de 4 dígitos en línea con USME
        if "USME" in line.upper():
            match = re.search(r"\b(\d{4})\b", line)
            if match:
                numero_proyecto = match.group(1)

        # Solicitud No.
        if "SOLICITUD NO" in line.upper():
            num_match = re.search(r"(\d+)", line)
            if num_match:
                numero_oficio = num_match.group(1)

        # Fecha CDP
        if "CDP DE FECHA" in line.upper():
            fecha_match = re.search(r"(\d{4})/(\d{2})/(\d{2})", line)
            if fecha_match:
                fecha_oficio = f"{fecha_match.group(3)}/{fecha_match.group(2)}/{fecha_match.group(1)}"

    pep_convertido = convertir_pep(numero_proyecto if numero_proyecto else "0000")

    registro = {
        "Archivo": os.path.basename(pdf_path),
        "importe Original": valor,
        "Posición Presupuestal": "10",
        "Elemento PEP": pep_convertido,
        "Objeto": objeto,
        "Número Oficio": numero_oficio,
        "Fecha Oficio": fecha_oficio
    }

    log_lines.append({"Archivo": os.path.basename(pdf_path), "Estado": f"✔️ Proyecto {numero_proyecto} → {pep_convertido}, Valor {valor}"})
    return [registro]

# 📂 Ruta fija
carpeta = r"C:\RICHARD\FDL\Usme\2026\CDP_Vigencia\Enero\Quinto_Grupo"
archivos = [os.path.join(carpeta, f) for f in os.listdir(carpeta) if f.lower().endswith(".pdf")]

todos = []
log_lines = []

for archivo in archivos:
    todos.extend(extraer_cdps(archivo, log_lines))

# Generar Excel
df = pd.DataFrame(todos)

df["CDP"] = range(1, len(df) + 1)
df["Num. Ext. Entidad"] = range(1, len(df) + 1)

fecha_actual = datetime.today().strftime("%d.%m.%Y")
fijos = {
    "Posición": "1",
    "Clase Documento": "CP",
    "Sociedad": "1001",
    "Moneda": "COP",
    "Fondos": "1-100-I079",
    "Periodo Presupuestario": "2026",
    "Cuenta de Mayor": "7990990000",
    "ID Solicitante": "1000131265",
    "ID Responsable": "1000835316",
    "Fecha Documento": fecha_actual,
    "Fecha Contabilización": fecha_actual
}

for col, val in fijos.items():
    df[col] = val

columnas_finales = [
    "CDP", "Posición", "Fecha Documento", "Fecha Contabilización", "Clase Documento",
    "Sociedad", "Moneda", "importe Original", "Posición Presupuestal", "Fondos",
    "Elemento PEP", "Periodo Presupuestario", "Cuenta de Mayor", "Objeto",
    "Número Oficio", "Fecha Oficio", "ID Solicitante", "ID Responsable",
    "Num. Ext. Entidad", "Archivo"
]

df_final = df[columnas_finales]
salida = os.path.join(carpeta, "Plantilla_CDP_Generada.xlsx")
df_final.to_excel(salida, index=False)

# Agregar hoja de log
libro = load_workbook(salida)
hoja_log = libro.create_sheet("Log_Auditoría")
hoja_log.append(["Archivo", "Estado"])
for fila in log_lines:
    hoja_log.append([fila["Archivo"], fila["Estado"]])
hoja_log.append([])
hoja_log.append(["Total PDFs procesados", len(archivos)])
hoja_log.append(["Registros exportados", len(df_final)])
libro.save(salida)

print(f"✅ Plantilla generada en: {salida} con hoja de auditoría incluida.")








