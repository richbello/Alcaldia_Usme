import re
import fitz  # PyMuPDF
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

def limpiar_numero(s: str) -> int:
    if not s or s in ["-", ""]:
        return 0
    s = str(s).replace("$", "").replace(" ", "").strip()
    s = s.replace(".", "").replace(",", "")
    return int(s) if s.isdigit() else 0

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    texto = str(texto).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto

def convertir_pep(numero: str) -> str:
    return f"PM/0005/0101/45990002{numero}"

def extraer_cdps(pdf_path, log_lines):
    doc = fitz.open(pdf_path)
    pep_encontrado = None
    ultimo_registro = None

    for page in doc:
        texto = page.get_text("text")

        # Buscar todos los posibles números PEP (último válido)
        pep_matches = re.findall(r"\b(\d{4})\b", texto)
        for match in pep_matches:
            if match.isdigit():
                pep_encontrado = match  # se actualiza con el último encontrado

        objeto_match = re.search(r"OBJETO:\s*(.+?)(?=VALOR:)", texto, re.S)
        objeto = normalizar_texto(objeto_match.group(1)) if objeto_match else ""

        valor_match = re.search(r"VALOR:\s*\$?\s*([\d\.,]+)", texto)
        valor = limpiar_numero(valor_match.group(1)) if valor_match else 0

        solicitud_match = re.search(r"PARA LA SOLICITUD No\.?\s*(\d+)", texto)
        numero_oficio = solicitud_match.group(1) if solicitud_match else "No encontrado"

        fecha_match = re.search(r"CDP DE FECHA (\d{4})/(\d{2})/(\d{2})", texto)
        fecha_oficio = f"{fecha_match.group(3)}/{fecha_match.group(2)}/{fecha_match.group(1)}" if fecha_match else datetime.today().strftime("%d/%m/%Y")

        if valor > 0 and pep_encontrado:
            pep_convertido = convertir_pep(pep_encontrado)
            ultimo_registro = {
                "importe Original": valor,
                "Posición Presupuestal": "10",
                "Elemento PEP": pep_convertido,
                "Objeto": objeto,
                "Número Oficio": numero_oficio,
                "Fecha Oficio": fecha_oficio
            }

    if ultimo_registro:
        log_lines.append({"Archivo": os.path.basename(pdf_path), "Estado": "✔️ Registro válido"})
        return [ultimo_registro]
    else:
        log_lines.append({"Archivo": os.path.basename(pdf_path), "Estado": "❌ Sin datos válidos (valor o PEP faltante)"})
        return []

# 📂 Ruta fija
carpeta = r"C:\RICHARD\FDL\Usme\2026\CDP_Vigencia\Enero\Quinto_Grupo"
archivos = [os.path.join(carpeta, f) for f in os.listdir(carpeta) if f.lower().endswith(".pdf")]

todos = []
log_lines = []

for archivo in archivos:
    todos.extend(extraer_cdps(archivo, log_lines))

# Generar Excel
if todos:
    df = pd.DataFrame(todos)
    df = df[df["importe Original"] > 0]

    df["CDP"] = range(1, len(df) + 1)
    df["Num. Ext. Entidad"] = range(1, len(df) + 1)

    fecha_actual = datetime.today().strftime("%d.%m.%Y")
    fijos = {
        "Posición": "1",
        "Clase Documento": "CP",
        "Sociedad": "1001",
        "Moneda": "COP",
        "Fondos": "1-100-I079",
        "Periodo Presupuestario": "2026",
        "Cuenta de Mayor": "7990990000",
        "ID Solicitante": "1000131265",
        "ID Responsable": "1000835316",
        "Fecha Documento": fecha_actual,
        "Fecha Contabilización": fecha_actual
    }

    for col, val in fijos.items():
        df[col] = val

    columnas_finales = [
        "CDP", "Posición", "Fecha Documento", "Fecha Contabilización", "Clase Documento",
        "Sociedad", "Moneda", "importe Original", "Posición Presupuestal", "Fondos",
        "Elemento PEP", "Periodo Presupuestario", "Cuenta de Mayor", "Objeto",
        "Número Oficio", "Fecha Oficio", "ID Solicitante", "ID Responsable",
        "Num. Ext. Entidad"
    ]

    df_final = df[columnas_finales]
    salida = os.path.join(carpeta, "Plantilla_CDP_Generada.xlsx")
    df_final.to_excel(salida, index=False)

    # Agregar hoja de log
    libro = load_workbook(salida)
    hoja_log = libro.create_sheet("Log_Auditoría")
    hoja_log.append(["Archivo", "Estado"])
    for fila in log_lines:
        hoja_log.append([fila["Archivo"], fila["Estado"]])
    hoja_log.append([])
    hoja_log.append(["Total PDFs procesados", len(archivos)])
    hoja_log.append(["Registros válidos exportados", len(df_final)])
    libro.save(salida)

    print(f"✅ Plantilla generada en: {salida} con hoja de auditoría incluida.")
else:
    print("⚠️ No se encontraron registros válidos en los PDFs.")
















